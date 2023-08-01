# 姓名：郭宏亮
# 时间：2023/7/31 21:18
# pip install openpyxl
# 读取excel，实现文件驱动自动化执行
import openpyxl

from test_excel.VAR import EXCEL_PATH, EXCEL_PATH2, EXCEL_PATH3


def read_excel():
    excel = openpyxl.load_workbook(EXCEL_PATH)
    sheet = excel["Sheet1"]
    # print(sheet)
    # 获取页签的内容对象
    # print(sheet.values)
    # 创建装载excel数据的list
    list_tuple = []
    for value in sheet.values:
        # 打印每行的内容
        # value是自动以的变量名
        # 每次循环都是一行内容
        # print(value)
        if type(value[0]) is int:
            list_tuple.append(value)

    # print(list_tuple)

    return list_tuple


def read_excel2():
    excel = openpyxl.load_workbook(EXCEL_PATH2)
    sheet = excel["Sheet1"]
    # print(sheet)
    # 获取页签的内容对象
    # print(sheet.values)
    # 创建装载excel数据的list
    list_tuple = []
    for value in sheet.values:
        # 打印每行的内容
        # value是自动以的变量名
        # 每次循环都是一行内容
        # print(value)
        if type(value[0]) is int:
            list_tuple.append(value)

    # print(list_tuple)

    return list_tuple


def read_excel3():
    excel = openpyxl.load_workbook(EXCEL_PATH3)
    sheet = excel["Sheet1"]
    # print(sheet)
    # 获取页签的内容对象
    # print(sheet.values)
    # 创建装载excel数据的list
    list_tuple = []
    for value in sheet.values:
        # 打印每行的内容
        # value是自动以的变量名
        # 每次循环都是一行内容
        # print(value)
        if type(value[0]) is int:
            list_tuple.append(value)

    # print(list_tuple)

    return list_tuple


if __name__ == '__main__':
    print(read_excel())
