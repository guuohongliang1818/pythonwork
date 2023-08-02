# 姓名：郭宏亮
# 时间：2023/7/31 21:41
import allure
import openpyxl
import pytest

from test_yaml.VAR import EXCEL_PATH, EXCEL_PATH2
from test_yaml.api_keyword.api_key import ApiKey
from test_yaml.data_driver.excel_read import read_excel, read_excel2


# 在当前文件中，所有的用例执行之前运行一次

# @pytest.fixture(scope="class")
def setup_module():
    # 1.定义全局变量
    global ak, excel, sheet
    # 实例化工具类
    ak = ApiKey()
    # 初始化excel文件
    excel = openpyxl.load_workbook(EXCEL_PATH2)
    sheet = excel["Sheet1"]


@pytest.mark.parametrize("data", read_excel2())
def test_01(data):
    # 动态生成用例标题
    if data[11] is not None:
        allure.dynamic.title(data[11])

    if data[16] is not None:
        allure.dynamic.story(data[16])

    if data[17] is not None:
        allure.dynamic.feature(data[17])

    if data[18] is not None:
        allure.dynamic.description(data[18])

    if data[19] is not None:
        allure.dynamic.severity(data[19])
    # ==========excel数据解析==========
    # print(data)
    r = data[0] + 1
    try:
        dict_data = {
            "url": data[1] + data[2],
            "params": eval(data[4]),
            "headers": eval(data[5]),
            data[7]: eval(data[6])
        }
    except:
        print("========实际结果=======")
        print("请求参数有误，请检查")

        # 这里列是从1开始的
        sheet.cell(r, 11).value = "请求参数有误，请检查"
        excel.save(EXCEL_PATH2)

    # 发起请求
    print(dict_data)
    # res = ak.post(url=dict_data.get("url"), params=dict_data.get("params"), json=dict_data.get(data[7]))
    res = ak.post(**dict_data)
    print(res.text)
    print(type(res.text))
    print(type(res.json()))

    # 结果检查
    # 实际结果
    try:
        result = None
        result = ak.get_text(res.text, data[8])
        print(result == data[9])
        if result == data[9]:
            sheet.cell(r, 11).value = "通过"
        else:
            sheet.cell(r, 11).value = "不通过"
        excel.save(EXCEL_PATH2)
    except:
        print("==============实际结果==============")
        print("预期结果的jsonpath表达式有误，请检查")
        sheet.cell(r, 11).value = "预期结果的jsonpath表达式有误，请检查"
        excel.save(EXCEL_PATH2)
    finally:
        assert result == data[9]


# 反射
# res = getattr(ak, data[3])(**dict_data)


if __name__ == '__main__':
    pytest.main(["-s", "test_excel_v1.py::test_01"])
